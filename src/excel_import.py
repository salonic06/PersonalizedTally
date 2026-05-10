from __future__ import annotations

import re
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Iterable, Optional

from openpyxl import load_workbook

from .repo import normalize_customer_name


@dataclass(frozen=True)
class ImportedInvoice:
    invoice_no: str
    invoice_date: date
    customer_name: str
    customer_address: str
    customer_gstin: str
    customer_state: str
    customer_state_code: str
    ship_name: str
    ship_address: str
    ship_gstin: str
    ship_state: str
    ship_state_code: str
    total_after_tax: float
    excel_path: str
    lines: list[dict]


_RE_AFTER_COLON = re.compile(r"^\s*[^:]+:\s*(.+?)\s*$")
_RE_ADDR_PREFIX = re.compile(r"^\s*Address\s*:\s*", re.IGNORECASE)


def _normalize_address_cell(value: object) -> str:
    if value is None:
        return ""
    s = str(value).strip()
    if not s:
        return ""
    m = _RE_ADDR_PREFIX.match(s)
    if m:
        return s[m.end() :].strip()
    return _parse_after_colon(s) or s


def _parse_after_colon(value: object) -> Optional[str]:
    if value is None:
        return None
    s = str(value).strip()
    if not s:
        return None
    m = _RE_AFTER_COLON.match(s)
    return (m.group(1).strip() if m else s) or None


def _parse_invoice_date(value: object) -> Optional[date]:
    if value is None:
        return None
    # Template cell may be: "Tax Invoice Dt.: 20.02.2026"
    s = str(value).strip()
    if not s:
        return None
    s2 = _parse_after_colon(s) or s

    # Try common formats.
    for fmt in ("%d.%m.%Y", "%d-%m-%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(s2, fmt).date()
        except ValueError:
            pass
    return None


def _parse_amount(value: object) -> Optional[float]:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except Exception:
        try:
            return float(str(value).replace(",", "").strip())
        except Exception:
            return None


def iter_invoice_excels(folder: Path) -> Iterable[Path]:
    for p in folder.rglob("*.xlsx"):
        name = p.name
        # Ignore Excel temp/lock files.
        if name.startswith("~$"):
            continue
        yield p


def read_invoice_from_excel(path: Path) -> ImportedInvoice:
    # data_only=True reads cached formula results saved by Excel.
    wb = load_workbook(path, data_only=True, read_only=True)
    ws = wb.active

    inv_no_raw = ws["A10"].value
    inv_dt_raw = ws["A11"].value
    cust_raw = ws["A16"].value
    addr_raw = ws["A17"].value
    gst_raw = ws["A21"].value
    state_raw = ws["A22"].value
    state_code_raw = ws["H22"].value
    ship_name_raw = ws["I21"].value
    ship_addr_raw = ws["I22"].value
    ship_gst_raw = ws["P22"].value
    ship_state_raw = ws["I23"].value
    ship_state_code_raw = ws["P23"].value
    total_raw = ws["O43"].value

    invoice_no = _parse_after_colon(inv_no_raw) or ""
    invoice_date = _parse_invoice_date(inv_dt_raw)
    customer_name_raw = str(cust_raw).strip() if cust_raw is not None else ""
    # Some templates store values like "Name: ABC Company" in A16.
    # Extract after ":" when present.
    customer_name = _parse_after_colon(customer_name_raw) or customer_name_raw
    customer_name = normalize_customer_name(customer_name)
    customer_address = _normalize_address_cell(addr_raw)
    customer_gstin = _parse_after_colon(gst_raw) or (str(gst_raw).strip() if gst_raw is not None else "")
    customer_state = _parse_after_colon(state_raw) or (str(state_raw).strip() if state_raw is not None else "")
    _h22 = str(state_code_raw).strip() if state_code_raw is not None else ""
    customer_state_code = _parse_after_colon(_h22) or _h22

    ship_name = _parse_after_colon(ship_name_raw) or (str(ship_name_raw).strip() if ship_name_raw is not None else "")
    ship_address = _normalize_address_cell(ship_addr_raw)
    ship_gstin = _parse_after_colon(ship_gst_raw) or (str(ship_gst_raw).strip() if ship_gst_raw is not None else "")
    ship_state = _parse_after_colon(ship_state_raw) or (str(ship_state_raw).strip() if ship_state_raw is not None else "")
    ship_state_code = _parse_after_colon(ship_state_code_raw) or (
        str(ship_state_code_raw).strip() if ship_state_code_raw is not None else ""
    )
    total_after_tax = _parse_amount(total_raw)

    if not invoice_no:
        raise ValueError(f"Missing invoice number in A10 for {path.name}")
    if not customer_name:
        raise ValueError(f"Missing customer name in A16 for {path.name}")
    if invoice_date is None:
        raise ValueError(f"Could not parse invoice date in A11 for {path.name}")
    if total_after_tax is None:
        raise ValueError(f"Missing/invalid total in O43 for {path.name}")

    # Line items (max 7): try table layout first (rows 25..31).
    lines: list[dict] = []
    for i in range(7):
        r = 25 + i
        desc = ws[f"B{r}"].value
        if desc is None or str(desc).strip() == "":
            continue
        lines.append(
            {
                "line_no": i + 1,
                "description": str(desc).strip(),
                "hsn": str(ws[f"I{r}"].value or "").strip(),
                "qty": _parse_amount(ws[f"K{r}"].value),
                "unit": str(ws[f"L{r}"].value or "").strip(),
                "rate": _parse_amount(ws[f"M{r}"].value),
                "amount": _parse_amount(ws[f"O{r}"].value),
            }
        )

    # Fallback for older single-line template (common cells from earlier automation).
    if not lines:
        desc2 = ws["A21"].value
        if desc2 is not None and str(desc2).strip():
            lines.append(
                {
                    "line_no": 1,
                    "description": str(desc2).strip(),
                    "hsn": str(ws["A22"].value or "").strip(),
                    "qty": _parse_amount(ws["H22"].value),
                    "unit": "Kgs",
                    "rate": _parse_amount(ws["I22"].value),
                    "amount": _parse_amount(ws["P22"].value),
                }
            )

    return ImportedInvoice(
        invoice_no=invoice_no,
        invoice_date=invoice_date,
        customer_name=customer_name,
        customer_address=customer_address,
        customer_gstin=customer_gstin,
        customer_state=customer_state,
        customer_state_code=customer_state_code,
        ship_name=ship_name,
        ship_address=ship_address,
        ship_gstin=ship_gstin,
        ship_state=ship_state,
        ship_state_code=ship_state_code,
        total_after_tax=float(total_after_tax),
        excel_path=str(path),
        lines=lines,
    )

