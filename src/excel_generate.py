from __future__ import annotations

import os
import re
import shutil
from copy import copy
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

# Display prefix on invoice (import/normalize may strip M/s elsewhere).
_RE_MS_PREFIX = re.compile(r"^\s*M\s*/\s*S\.?\s*", re.IGNORECASE)

# Outer invoice frame: A1 through P52 (template may merge down from J47 across to P52).
# Applied only in openpyxl post-process, *after* all cell value/font writes, so edits do not strip it.
_BORDER_MIN_ROW = 1
_BORDER_MAX_ROW = 52
_BORDER_MIN_COL = 1
_BORDER_MAX_COL = 16  # column P
_OUTER_BORDER_SIDE_STYLE = "medium"

# Item block in template (must match `generate_invoice_excel`).
_ITEM_TABLE_START_ROW = 25
_ITEM_TABLE_MAX_ROWS = 7
# Split long product+note text across extra Excel rows (column B only) instead of one tall wrapped cell.
_ITEM_DESC_CHUNK_CHARS = 52

# Grand total cell (must match `excel_import.read_invoice_from_excel`).
_GRAND_TOTAL_CELL = "O43"

# Merged cells A13 / A44: same bold-label + normal-value pattern as A11, I10, etc.
_INVOICE_A13_LABEL = "State: "
_INVOICE_A13_VALUE = "Maharashtra"
_INVOICE_A44_LABEL = "Bank Details: \n"
_INVOICE_A44_VALUE = (
    "Bank Name & Branch : Canara Bank, Bhavani Peth Branch,\n"
    "Bank A/C: 0361261004246 - IFSC: CNRB0000361"
)

@dataclass(frozen=True)
class InvoiceLine:
    description: str
    hsn: str
    qty: float
    unit: str
    rate: float


@dataclass(frozen=True)
class GenerateInvoiceInput:
    template_path: Path
    output_path: Path
    invoice_no: str
    invoice_date_ddmmyyyy: str
    customer_name: str
    customer_address: str = ""
    customer_gstin: str = ""
    customer_state: str = ""
    customer_state_code: str = ""
    transport_mode: str = ""
    eway_bill_no: str = ""
    payment_terms: str = ""
    place_of_supply: str = ""
    po_no: str = ""
    po_date: str = ""
    shipped_same_as_billed: bool = True
    ship_name: str = ""
    ship_address: str = ""
    ship_gstin: str = ""
    ship_state: str = ""
    ship_state_code: str = ""
    lines: list[InvoiceLine] = field(default_factory=list)


def compute_gst_invoice_totals(
    lines: Iterable[InvoiceLine],
    *,
    gst_rate: float = 0.18,
) -> tuple[float, float, float]:
    """
    Taxable line subtotal (qty × rate), GST, and grand total.
    Same rules as the invoice preview UI (flat rate on subtotal).
    """
    subtotal = 0.0
    for line in lines:
        subtotal += float(line.qty) * float(line.rate)
    subtotal = round(subtotal, 2)
    gst = round(subtotal * float(gst_rate), 2)
    total = round(subtotal + gst, 2)
    return subtotal, gst, total


def _number_to_words(n: float) -> str:
    """
    Amount in words for invoices. Whole rupees use Indian numbering; fractional paise
    uses each decimal digit as a word (e.g. 5.23 -> Rupees Five and Two Three Only).
    """
    units = [
        "",
        "One",
        "Two",
        "Three",
        "Four",
        "Five",
        "Six",
        "Seven",
        "Eight",
        "Nine",
        "Ten",
        "Eleven",
        "Twelve",
        "Thirteen",
        "Fourteen",
        "Fifteen",
        "Sixteen",
        "Seventeen",
        "Eighteen",
        "Nineteen",
    ]
    tens = ["", "", "Twenty", "Thirty", "Forty", "Fifty", "Sixty", "Seventy", "Eighty", "Ninety"]

    def words(num: int) -> str:
        if num == 0:
            return ""
        if num < 20:
            return units[num]
        if num < 100:
            return tens[num // 10] + (" " + units[num % 10] if num % 10 else "")
        if num < 1000:
            return units[num // 100] + " Hundred" + (" " + words(num % 100) if num % 100 else "")
        if num < 100000:
            return words(num // 1000) + " Thousand" + (" " + words(num % 1000) if num % 1000 else "")
        if num < 10000000:
            return words(num // 100000) + " Lakh" + (" " + words(num % 100000) if num % 100000 else "")
        return words(num // 10000000) + " Crore" + (" " + words(num % 10000000) if num % 10000000 else "")

    digit_words = [
        "Zero",
        "One",
        "Two",
        "Three",
        "Four",
        "Five",
        "Six",
        "Seven",
        "Eight",
        "Nine",
    ]

    rounded = round(float(n) + 1e-9, 2)
    s = f"{rounded:.2f}"
    int_part_str, frac_part = s.split(".", 1)
    int_part = int(int_part_str)
    frac_digits = "" if int(frac_part) == 0 else frac_part

    int_words = words(int_part) if int_part > 0 else "Zero"
    if not frac_digits:
        return f"Rupees {int_words} Only"

    frac_spelled = " ".join(digit_words[int(ch)] for ch in frac_digits if ch.isdigit())
    return f"Rupees {int_words} and {frac_spelled} Only"


def _cell_str(value: str) -> str:
    return (value or "").strip()


def _with_ms_prefix(name: str) -> str:
    """Invoice display: ensure customer/ship name starts with 'M/s.' when non-empty."""
    s = _cell_str(name)
    if not s:
        return ""
    if _RE_MS_PREFIX.match(s):
        return s
    return f"M/s. {s}"


def format_invoice_customer_display_name(name: str) -> str:
    """M/s. prefix for UI preview — same rule as Excel output."""
    return _with_ms_prefix(name)


def _label(prefix: str, value: str) -> str:
    """Rebuild template-style 'Label: value' lines (import strips these prefixes)."""
    return f"{prefix}{_cell_str(value)}"


def _anchor_oxl(ws: Worksheet, coordinate: str):  # type: ignore[name-defined]
    """Top-left writable cell when `coordinate` lies inside a merged range."""
    from openpyxl.cell.cell import MergedCell

    c = ws[coordinate]
    if isinstance(c, MergedCell):
        for rng in ws.merged_cells.ranges:
            if coordinate in rng:
                return ws.cell(rng.min_row, rng.min_col)
        return c
    for rng in ws.merged_cells.ranges:
        if coordinate in rng:
            return ws.cell(rng.min_row, rng.min_col)
    return c


def _no_bold_oxl(cell) -> None:  # type: ignore[no-untyped-def]
    from openpyxl.styles import Font

    f = cell.font
    if f is not None:
        nf = copy(f)
        nf.bold = False
        cell.font = nf
    else:
        cell.font = Font(bold=False)


def _assign_oxl(ws: Worksheet, cell: str, value: object) -> None:  # type: ignore[name-defined]
    tl = _anchor_oxl(ws, cell)
    tl.value = value
    _no_bold_oxl(tl)


def _assign_flat_for_richtext_oxl(
    ws: Worksheet,  # type: ignore[name-defined]
    cell: str,
    label: str,
    value: str,
    rich_queue: list[tuple[str, str, str]],
) -> None:
    val = _cell_str(value)
    rich_queue.append((cell, label, val))
    _assign_oxl(ws, cell, f"{label}{val}")


def _assign_full_bold_oxl(
    ws: Worksheet,  # type: ignore[name-defined]
    cell: str,
    text: str,
    full_bold: set[str],
) -> None:
    _assign_oxl(ws, cell, text)
    full_bold.add(cell)


def _safe_clear_contents_oxl(ws: Worksheet, cell: str) -> None:  # type: ignore[name-defined]
    tl = _anchor_oxl(ws, cell)
    tl.value = None


def _apply_invoice_outer_border(ws: Worksheet) -> None:  # type: ignore[name-defined]
    """Single rectangular frame A1:P52; inner grid edges kept; outer sides use _OUTER_BORDER_SIDE_STYLE."""
    from openpyxl.styles import Border, Side

    outer = Side(style=_OUTER_BORDER_SIDE_STYLE, color="000000")
    r1, r2 = _BORDER_MIN_ROW, _BORDER_MAX_ROW
    c1, c2 = _BORDER_MIN_COL, _BORDER_MAX_COL
    for row in range(r1, r2 + 1):
        for col in range(c1, c2 + 1):
            cell = ws.cell(row=row, column=col)
            prev = cell.border
            left = outer if col == c1 else (prev.left if prev and prev.left else None)
            right = outer if col == c2 else (prev.right if prev and prev.right else None)
            top = outer if row == r1 else (prev.top if prev and prev.top else None)
            bottom = outer if row == r2 else (prev.bottom if prev and prev.bottom else None)
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def _postprocess_invoice_xlsx(
    path: str,
    *,
    split_entries: list[tuple[str, str, str]],
    full_bold_cells: set[str],
) -> None:
    """
    After base save: (1) rich-text (bold label + normal value) and full-bold cells;
    (2) redraw the outer frame last; (3) save.
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.cell.rich_text import CellRichText, TextBlock
        from openpyxl.cell.text import InlineFont
        from openpyxl.styles import Font
    except ImportError:
        return

    try:
        wb = load_workbook(path)
        ws = wb.active

        for addr, label, value in split_entries:
            cell = ws[addr]
            base = cell.font
            name = (base.name if base and base.name else None) or "Calibri"
            sz = float(base.sz) if base and base.sz is not None else 11.0
            bold_if = InlineFont(rFont=name, sz=sz, b=True)
            norm_if = InlineFont(rFont=name, sz=sz, b=False)
            cell.value = CellRichText(
                TextBlock(bold_if, label),
                TextBlock(norm_if, value or ""),
            )

        for addr in full_bold_cells:
            cell = ws[addr]
            f = cell.font
            if f is not None:
                nf = copy(f)
                nf.bold = True
                cell.font = nf
            else:
                cell.font = Font(bold=True)

        _apply_invoice_outer_border(ws)
        wb.save(path)
    except Exception:
        pass


def _split_description_chunks(text: str, *, max_chars: int, max_chunks: int) -> list[str]:
    """
    Word-wrap description into up to `max_chunks` segments of at most `max_chars` characters.
    Used to place overflow on the next Excel row (column B only) instead of one very tall cell.
    """
    s = (text or "").replace("\r\n", "\n").strip()
    if not s:
        return [""]
    words: list[str] = []
    for ln in s.split("\n"):
        ln = ln.strip()
        if ln:
            words.extend(ln.split())
    if not words:
        return [""]
    chunks: list[str] = []
    i = 0
    while i < len(words) and len(chunks) < max_chunks:
        cur: list[str] = []
        nlen = 0
        while i < len(words):
            w = words[i]
            extra = len(w) if not cur else 1 + len(w)
            if nlen + extra > max_chars:
                break
            cur.append(w)
            nlen += extra
            i += 1
        if cur:
            chunks.append(" ".join(cur))
            continue
        w = words[i]
        if len(w) <= max_chars:
            chunks.append(w)
        else:
            chunks.append(w[: max(1, max_chars - 1)] + "…")
        i += 1
    if i < len(words) and chunks:
        tail = chunks[-1]
        if not tail.endswith("…"):
            if len(tail) > max_chars - 1:
                cut = tail[: max_chars - 1].rsplit(" ", 1)[0].rstrip()
                tail = cut if cut else tail[: max_chars - 1]
            chunks[-1] = tail + "…"
    return chunks


def _plan_item_excel_rows(
    lines: list[InvoiceLine],
    *,
    start_row: int,
    max_rows: int,
    chunk_chars: int,
) -> list[tuple[int, int | None, str, InvoiceLine | None]]:
    """
    Build physical Excel rows for the item block.
    Returns tuples: (row, serial or None for continuation, B-column text, line or None).
    Amounts/HSN/qty are written only when `line` is not None.
    """
    out: list[tuple[int, int | None, str, InvoiceLine | None]] = []
    row_off = 0
    serial = 0
    for line in lines:
        space = max_rows - row_off
        if space <= 0:
            raise ValueError(
                "Invoice item area is full (7 rows in the template). "
                "Use fewer product lines or shorten descriptions / notes."
            )
        parts = _split_description_chunks(
            line.description, max_chars=chunk_chars, max_chunks=space
        )
        for j, btext in enumerate(parts):
            r = start_row + row_off
            if j == 0:
                serial += 1
                out.append((r, serial, btext, line))
            else:
                out.append((r, None, btext, None))
            row_off += 1
    return out


def generate_invoice_excel(inp: GenerateInvoiceInput) -> tuple[float, str]:
    """
    Build the invoice `.xlsx` from the user's template using **openpyxl only**
    (no Microsoft Excel / COM). Line amounts and grand total are written explicitly so
    `data_only` imports see correct values without an Excel recalc pass.

    Returns: (total_after_tax, saved_path)
    """
    from openpyxl import load_workbook

    lines = list(inp.lines or [])[:_ITEM_TABLE_MAX_ROWS]

    template = os.path.normpath(os.path.abspath(str(inp.template_path)))
    out_path = os.path.normpath(os.path.abspath(str(inp.output_path)))
    if not os.path.isfile(template):
        raise FileNotFoundError(f"Invoice template not found: {template}")

    os.makedirs(os.path.dirname(out_path) or ".", exist_ok=True)
    if os.path.isfile(out_path):
        try:
            os.remove(out_path)
        except OSError as e:
            raise OSError(
                f"Cannot overwrite existing file (close it if open in another program): {out_path}"
            ) from e

    shutil.copy2(template, out_path)

    wb = load_workbook(out_path)
    ws = wb.active

    rich_queue: list[tuple[str, str, str]] = []
    full_bold: set[str] = set()

    bill_name = _with_ms_prefix(inp.customer_name)
    ship_name = bill_name if inp.shipped_same_as_billed else _with_ms_prefix(inp.ship_name)

    _assign_full_bold_oxl(ws, "A10", f"Tax Invoice No: {_cell_str(inp.invoice_no)}", full_bold)
    _assign_flat_for_richtext_oxl(
        ws, "A11", "Tax Invoice Dt.: ", _cell_str(inp.invoice_date_ddmmyyyy), rich_queue
    )
    _assign_flat_for_richtext_oxl(ws, "I10", "Transport Mode: ", inp.transport_mode, rich_queue)
    _assign_flat_for_richtext_oxl(ws, "I11", "Eway Bill No: ", inp.eway_bill_no, rich_queue)
    _assign_flat_for_richtext_oxl(ws, "I12", "Payment Terms: ", inp.payment_terms, rich_queue)
    _assign_flat_for_richtext_oxl(ws, "I13", "Place of Supply: ", inp.place_of_supply, rich_queue)
    _assign_oxl(ws, "C14", _cell_str(inp.po_no))
    _assign_oxl(ws, "K14", _cell_str(inp.po_date))
    _assign_flat_for_richtext_oxl(ws, "A13", _INVOICE_A13_LABEL, _INVOICE_A13_VALUE, rich_queue)
    _assign_flat_for_richtext_oxl(ws, "A44", _INVOICE_A44_LABEL, _INVOICE_A44_VALUE, rich_queue)

    _assign_full_bold_oxl(ws, "A16", f"Name: {bill_name}", full_bold)
    _assign_flat_for_richtext_oxl(ws, "A17", "Address: ", inp.customer_address, rich_queue)
    _assign_flat_for_richtext_oxl(ws, "A21", "GSTIN: ", inp.customer_gstin, rich_queue)
    _assign_full_bold_oxl(ws, "A22", f"State: {_cell_str(inp.customer_state)}", full_bold)
    _assign_full_bold_oxl(ws, "H22", _cell_str(inp.customer_state_code), full_bold)

    if inp.shipped_same_as_billed:
        _assign_full_bold_oxl(ws, "I16", f"Name: {ship_name}", full_bold)
        _assign_flat_for_richtext_oxl(ws, "I17", "Address: ", inp.customer_address, rich_queue)
        _assign_flat_for_richtext_oxl(ws, "I21", "GSTIN: ", inp.customer_gstin, rich_queue)
        _assign_full_bold_oxl(ws, "I22", f"State: {_cell_str(inp.customer_state)}", full_bold)
        _assign_full_bold_oxl(ws, "P22", _cell_str(inp.customer_state_code), full_bold)
    else:
        _assign_full_bold_oxl(ws, "I16", f"Name: {ship_name}", full_bold)
        _assign_flat_for_richtext_oxl(ws, "I17", "Address: ", inp.ship_address, rich_queue)
        _assign_flat_for_richtext_oxl(ws, "I21", "GSTIN: ", inp.ship_gstin, rich_queue)
        _assign_full_bold_oxl(ws, "I22", f"State: {_cell_str(inp.ship_state)}", full_bold)
        _assign_full_bold_oxl(ws, "P22", _cell_str(inp.ship_state_code), full_bold)

    start_row = _ITEM_TABLE_START_ROW
    max_rows = _ITEM_TABLE_MAX_ROWS
    _item_cols = ("A", "B", "I", "K", "L", "M", "O")
    for r in range(start_row, start_row + max_rows):
        for col in _item_cols:
            _safe_clear_contents_oxl(ws, f"{col}{r}")

    item_plan = _plan_item_excel_rows(
        lines,
        start_row=start_row,
        max_rows=max_rows,
        chunk_chars=_ITEM_DESC_CHUNK_CHARS,
    )
    for r, serial, btext, line in item_plan:
        if line is not None:
            line_total = float(line.qty) * float(line.rate)
            _assign_oxl(ws, f"A{r}", serial)
            _assign_oxl(ws, f"B{r}", _cell_str(btext))
            _assign_oxl(ws, f"I{r}", _cell_str(line.hsn))
            _assign_oxl(ws, f"K{r}", float(line.qty))
            _assign_oxl(ws, f"L{r}", _cell_str(line.unit) or "Nos")
            _assign_oxl(ws, f"M{r}", float(line.rate))
            _assign_oxl(ws, f"O{r}", line_total)
        else:
            _assign_oxl(ws, f"A{r}", "")
            _assign_oxl(ws, f"B{r}", _cell_str(btext))

    subtotal, _gst_amt, total_after_tax = compute_gst_invoice_totals(lines)
    if subtotal <= 0 and lines:
        raise ValueError("Invoice line subtotal is zero; check quantities and rates.")

    _anchor_oxl(ws, _GRAND_TOTAL_CELL).value = float(total_after_tax)

    _assign_flat_for_richtext_oxl(
        ws,
        "A39",
        "Total Amount in Words : ",
        _number_to_words(float(total_after_tax)),
        rich_queue,
    )

    wb.save(out_path)
    _postprocess_invoice_xlsx(out_path, split_entries=rich_queue, full_bold_cells=full_bold)
    return float(total_after_tax), out_path
